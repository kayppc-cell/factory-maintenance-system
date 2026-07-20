import streamlit as st
import requests
import json
import os
import shutil
import datetime
import qrcode
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import zipfile
import pandas as pd

# --- 1. CONFIGURATION ---
LINE_ACCESS_TOKEN = "SOs7DeGwVsFpuK/JN8zm58Wn3EOiB75Ww0q57z1/yht4H1imzYonre4QuPfQ3cxbJ7j9dpyNMSTviG06LCe//YM1+r5TqRQx09p8nLNh5lYwCp4biq7N20ffJqzGm+ZYNgtEzt2rYZ/GYVRV725EiAdB04t89/1O/w1cDnyilFU="
LINE_TARGET_ID = "Cbf3d27d5280ae8b258727047a26b399a"

BASE_FOLDER = os.path.dirname(os.path.abspath(__file__)) if "__file__" in locals() else os.getcwd()

BOSS_PASSWORD = "boss1234"
BIGBOSS_PASSWORD = "bigboss9999"

now = datetime.datetime.now()
current_time_str = now.strftime("%Y-%m-%d %H:%M:%S")

MACHINES = {
    "CNC3X-01": "CNC 3 แกน #01", "CNC3X-02": "CNC 3 แกน #02",
    "CNC3X-03": "CNC 3 แกน #03", "CNC3X-04": "CNC 3 แกน #04",
    "CNC3X-05": "CNC 3 แกน #05", "CNC3X-06": "CNC 3 แกน #06",
    "CNC3X-07": "CNC 3 แกน #07", "CNC3X-08": "CNC 3 แกน #08",
    "CNC5X-01": "CNC 5 แกน #พิเศษ",
    "Crane no.1": "เครน CNC NO.1", "Crane no.2": "เครน QC NO.2",
    "QC-01": "เครื่องวัดความแข็ง",
    "QC-02": "เวอร์เนีย 1000 QC-VN-009",
    "QC-03": "เวอร์เนีย 300 QC-VN-011",
    "QC-04": "เวอร์เนีย 300 QC-VN-029",
    "QC-05": "เวอร์เนีย 200 QC-VN-025",
    "QC-06": "เวอร์เนีย 200 QC-VN-026",
    "QC-07": "เวอร์เนีย 200 QC-VN-027",
    "QC-08": "เวอร์เนีย 200 QC-VN-028",
    "QC-09": "เวอร์เนีย 600 QC-VN-010",
    "QC-10": "ไฮเกจ 300 QC-HG-006",
    "QC-11": "ไฮเกจ 600 QC-HG-007",
    "QC-12": "ไฮเกจ 1000 QC-HG-008",
    "QC-13": "ไมโคร 0-25 QC-MC-013",
    "QC-14": "ไมโคร 5-30 QC-MC-024",
    "QC-15": "CMM QC-CMM-001",
    "QC-16": "Laser QC-Laser-001",
    "QC-17": "เลื่อยสายพาน QC-SAW-001",
    "QC-18": "Faro Arm QC-AC-001",
    "QC-19": "Cimcore Arm 2.8 QC-AC-003",
    "QC-20": "Cimcore Arm 2.4 QC-AC-002",
    "QC-21": "Cimcore Arm 3.5 QC-AC-004",
    "COMP-01": "ปั๊มลม 1 COMP-01",
    "COMP-02": "ปั๊มลม 2 COMP-02",
    "GRINDING-01": "เครื่องเจียร GRINDING #01", "GRINDING-02": "เครื่องเจียร GRINDING #02",
    "CUTTER GRINDING-01": "เครื่องลับคม CUTTER GRINDING #01",
    "MILLING-01": "เครื่องมิลลิ่ง #01", "MILLING-02": "เครื่องมิลลิ่ง #02", "MILLING-03": "เครื่องมิลลิ่ง #03", "MILLING-04": "เครื่องเฟสเก็บขนาด",
    "LATHE-01": "เครื่องกลึง LATHE #01",
    "CUTTING-01": "เครื่องตัด CUTTING #01",
    "BENDING-01": "เครื่องพับ BENDING #01",
    "MIG CO2-01": "เครื่องเชื่อม MIG CO2 #01", "MIG CO2-02": "เครื่องเชื่อม MIG CO2 #02", "MIG CO2-03": "เครื่องเชื่อม MIG CO2 #03",
    "ARGON-01": "เครื่องเชื่อม ARGON #01", "ARGON-02": "เครื่องเชื่อม ARGON #02",
    "WELDING_ALUMINUM-01": "เครื่องเชื่อมอลูมิเนียม WELDING ALUMINUM #01",
    "BAND SAW-01": "เครื่องเลื่อยสายพาน #01", "BAND SAW-02": "เครื่องเลื่อยสายพาน #02", "BAND SAW-03": "เครื่องเลื่อยสายพาน #03",
    "FORKLIFT-01": "รถโฟคลิฟ FORKLIFT #01"
}

CHECKLISTS = {
    "CNC": [
        "Worm up เครื่องจักร 15 นาที ทุกครั้งที่ใช้งาน", "เช็คระดับนำมัน Oil Matic Mesh ทุกวัน เติมเมื่อพร่อง",
        "ทำความสะอาด Air filter Mesh ทุกวัน", "ตรวจเช็คแรงดัน Air Control Unit ปกติเฉลี่ยที่ 0.5 Mpa",
        "เช็คน้ำมันชุด Gear ของ ATC ทุกวัน(เปลี่ยนถ่ายทุกปี)", "อัดจารบี Ballscrew และ Linear Guideทุก 1000 ชม.",
        "ตรวจสอบการเคลื่อนที่ของแกนทุกแกน (X,Y,Z)", "ตรวจสอบสภาพของน้ำ Coolant ถ่ายรูปค่าที่วัดได้ส่งเข้าระบบ",
        "การทำงานของ Coolant pump", "การทำงานของ Unclamp และการเปลี่ยน Tool", "การทำงานของ Spindle",
        "การทำงานของ Arm เปลี่ยน Tool", "ตรวจสอบระดับของน้ำ Coolant เติมเมื่ออยู่ระดับที่ต่ำ",
        "ความสะอาดทั่วไปของเครื่องจักรโดยรวม", "ตรวจสอบความพร้อมสภาพโดยรวม(ฟังด้วยหู ดูด้วยตา)", "ตรวจสอบสายไฮโดรลิกส์ และสายลม"
    ],
    "Crane no.1": [
        "ตรวจเช็คปุ่มกดต้องอยู่ในสภาพพร้อมใช้งาน ไม่แตก ไม่ชำรุด เสียหาย", "ตรวจเช็คการหยุดเครน เดินหน้า และถอยหลัง เมื่อปล่อยปุ่มกดต้องหยุดทันที",
        "ตรวจเช็คสลิงต้องไม่แตกฝอยเป็นหนาม\nและบิดงอ", "ตรวจเช็คตะขอต้องไม่มีรอยแตกร้าวสูญหายกิ๊ปปากตะขอไม่ชำรุดหรือหลุดหาย",
        "ตรวจเช็คสายบังคับเครนต้องไม่ชำรุดสายไฟไม่ขาดรุ่งริ่ง\nไม่เรียบร้อย", "ตรวจเช็คสัญญานเสียงเมื่อเริ่มเดินเครนต้องมีเสียงเตือนการทำงาน"
    ],
    "Crane no.2": [
        "ตรวจเช็คปุ่มกดต้องอยู่ในสภาพพร้อมใช้งาน ไม่แตก ไม่ชำรุด เสียหาย", "ตรวจเช็คการหยุดเครน เดินหน้า และถอยหลัง เมื่อปล่อยปุ่มกดต้องหยุดทันที",
        "ตรวจเช็คสลิงต้องไม่แตกฝอยเป็นหนาม\nและบิดงอ", "ตรวจเช็คตะขอต้องไม่มีรอยแตกร้าวสูญหายกิ๊ปปากตะขอไม่ชำรุดหรือหลุดหาย",
        "ตรวจเช็คสายบังคับเครนต้องไม่ชำรุดสายไฟไม่ขาดรุ่งริ่ง\nไม่เรียบร้อย", "ตรวจเช็คสัญญานเสียงเมื่อเริ่มเดินเครนต้องมีเสียงเตือนการทำงาน"
    ],
    "QC-01": ["ตรวจสอบความสะอาด", "ตรวจดู BATTERRY อ่อนหรือไม่", "ตรวจสอบปุ่มกดต่างๆๆ", "ตรวจสอบอุปกรณ์การชาร์จ"],
    "QC-02": ["ตรวจดูสภาพของเวอร์เนียพร้อมใช้งานหรือไม่", "ตรวจดู BATTERRY อ่อนหรือไม่", "ตรวจสอบการสไลด์ต้องไม่ติดขัด", "ตรวจสอบที่คีบตรงปลายที่ใช้วัดชิ้นงาน เช็คว่ามีรอยบิ่น หรือสึกหล่อหรือไม่", "หลังเลิกงานต้องปิดสวิตส์ทุกครั้ง"],
    "QC-03": ["ตรวจดูสภาพของเวอร์เนียพร้อมใช้งานหรือไม่", "ตรวจดู BATTERRY อ่อนหรือไม่", "ตรวจสอบการสไลด์ต้องไม่ติดขัด", "ตรวจสอบที่คีบตรงปลายที่ใช้วัดชิ้นงาน เช็คว่ามีรอยบิ่น หรือสึกหล่อหรือไม่", "หลังเลิกงานต้องปิดสวิตส์ทุกครั้ง"],
    "QC-04": ["ตรวจดูสภาพของเวอร์เนียพร้อมใช้งานหรือไม่", "ตรวจดู BATTERRY อ่อนหรือไม่", "ตรวจสอบการสไลด์ต้องไม่ติดขัด", "ตรวจสอบที่คีบตรงปลายที่ใช้วัดชิ้นงาน เช็คว่ามีรอยบิ่น หรือสึกหล่อหรือไม่", "หลังเลิกงานต้องปิดสวิตส์ทุกครั้ง"],
    "QC-05": ["ตรวจดูสภาพของเวอร์เนียพร้อมใช้งานหรือไม่", "ตรวจดู BATTERRY อ่อนหรือไม่", "ตรวจสอบการสไลด์ต้องไม่ติดขัด", "ตรวจสอบที่คีบตรงปลายที่ใช้วัดชิ้นงาน เช็คว่ามีรอยบิ่น หรือสึกหล่อหรือไม่", "หลังเลิกงานต้องปิดสวิตส์ทุกครั้ง"],
    "QC-06": ["ตรวจดูสภาพของเวอร์เนียพร้อมใช้งานหรือไม่", "ตรวจดู BATTERRY อ่อนหรือไม่", "ตรวจสอบการสไลด์ต้องไม่ติดขัด", "ตรวจสอบที่คีบตรงปลายที่ใช้วัดชิ้นงาน เช็คว่ามีรอยบิ่น หรือสึกหล่อหรือไม่", "หลังเลิกงานต้องปิดสวิตส์ทุกครั้ง"],
    "QC-07": ["ตรวจดูสภาพของเวอร์เนียพร้อมใช้งานหรือไม่", "ตรวจดู BATTERRY อ่อนหรือไม่", "ตรวจสอบการสไลด์ต้องไม่ติดขัด", "ตรวจสอบที่คีบตรงปลายที่ใช้วัดชิ้นงาน เช็คว่ามีรอยบิ่น หรือสึกหล่อหรือไม่", "หลังเลิกงานต้องปิดสวิตส์ทุกครั้ง"],
    "QC-08": ["ตรวจดูสภาพของเวอร์เนียพร้อมใช้งานหรือไม่", "ตรวจดู BATTERRY อ่อนหรือไม่", "ตรวจสอบการสไลด์ต้องไม่ติดขัด", "ตรวจสอบที่คีบตรงปลายที่ใช้วัดชิ้นงาน เช็คว่ามีรอยบิ่น หรือสึกหล่อหรือไม่", "หลังเลิกงานต้องปิดสวิตส์ทุกครั้ง"],
    "QC-09": ["ตรวจดูสภาพของเวอร์เนียพร้อมใช้งานหรือไม่", "ตรวจดู BATTERRY อ่อนหรือไม่", "ตรวจสอบการสไลด์ต้องไม่ติดขัด", "ตรวจสอบที่คีบตรงปลายที่ใช้วัดชิ้นงาน เช็คว่ามีรอยบิ่น หรือสึกหล่อหรือไม่", "หลังเลิกงานต้องปิดสวิตส์ทุกครั้ง"],
    "QC-10": ["ตรวจดูสภาพของไมโครพร้อมใช้งานหรือไม่", "ตรวจดู BATTERRY อ่อนหรือไม่", "ตรวจสอบการสไลด์ต้องไม่ติดขัด", "หลังเลิกงานต้องปิดสวิตส์ทุกครั้ง"],
    "QC-11": ["ตรวจดูสภาพของไมโครพร้อมใช้งานหรือไม่", "ตรวจดู BATTERRY อ่อนหรือไม่", "ตรวจสอบการสไลด์ต้องไม่ติดขัด", "หลังเลิกงานต้องปิดสวิตส์ทุกครั้ง"],
    "QC-12": ["ตรวจดูสภาพของไมโครพร้อมใช้งานหรือไม่", "ตรวจดู BATTERRY อ่อนหรือไม่", "ตรวจสอบการสไลด์ต้องไม่ติดขัด", "หลังเลิกงานต้องปิดสวิตส์ทุกครั้ง"],
    "QC-13": ["ตรวจดูสภาพของไมโครพร้อมใช้งานหรือไม่", "ตรวจดู BATTERRY อ่อนหรือไม่", "ตรวจสอบการหมุนเข้าหมุนออกต้องไม่ติดขัด", "หลังเลิกงานต้องปิดสวิตส์ทุกครั้ง"],
    "QC-14": ["ตรวจดูสภาพของไมโครพร้อมใช้งานหรือไม่", "ตรวจดู BATTERRY อ่อนหรือไม่", "ตรวจสอบการหมุนเข้าหมุนออกต้องไม่ติดขัด", "หลังเลิกงานต้องปิดสวิตส์ทุกครั้ง"],
    "QC-15": ["ตรวจดูสภาพของสายไฟ", "ตรวจดูสภาพของลมพร้อมใช้งานหรือไม่", "ตรวจดูสภาพของ SURFACE BASE", "ตรวจดูสภาพของ STICKER", "ตรวจสอบ COMPUTER", "ตรวจดูสภาพของหัว PROBE คตงอหรือไม่"],
    "QC-16": ["ตรวจดูสภาพของสายไฟ", "ตรวจดูสภาพ ของเครื่อง", "ตรวจดูสภาพของ Lend Laser", "ตรวจสอบ STICKER", "ตรวจสอบ COMPUTER"],
    "QC-17": ["ตรวจดูสภาพสายไฟ", "ตรวจดูสภาพของใบเลื่อย", "ตรวจดูสภาพของมอเตอร์", "ตรวจดูสภาพของสายพาน", "ตรวจสอบเครื่องเลื่อยสายพาน"],
    "QC-18": ["ตรวจดูสภาพของสายไฟ", "ตรวจดูสภาพ ARM ของเครื่อง", "ตรวจดูสภาพของหัว PROBE คตงอหรือไม่", "ตรวจสอบ STICKER", "ตรวจสอบ NOTEBOOK COMPUTER"],
    "QC-19": ["ตรวจดูสภาพของสายไฟ", "ตรวจดูสภาพ ARM ของเครื่อง", "ตรวจดูสภาพของหัว PROBE คตงอหรือไม่", "ตรวจสอบ STICKER", "ตรวจสอบ NOTEBOOK COMPUTER"],
    "QC-20": ["ตรวจดูสภาพของสายไฟ", "ตรวจดูสภาพ ARM ของเครื่อง", "ตรวจดูสภาพของหัว PROBE คตงอหรือไม่", "ตรวจสอบ STICKER", "ตรวจสอบ NOTEBOOK COMPUTER"],
    "QC-21": ["ตรวจดูสภาพของสายไฟ", "ตรวจดูสภาพ ARM ของเครื่อง", "ตรวจดูสภาพของหัว PROBE คตงอหรือไม่", "ตรวจสอบ STICKER", "ตรวจสอบ NOTEBOOK COMPUTER"],
    "COMP-01": ["เช็คแรงดัน (Pressure) ต้องไม่ต่ำกว่า 7 bar", "ตรวจสอบระดับน้ำมันไฮดรอลิก ต้องไม่ต่ำกว่าระดับต่ำสุด", "เช็คอุณหภูมิความร้อนต้องไม่เกิน 80 องศา", "เช็คการรั่วซีมของระบบน้ำมัน", "เช็คระบบเดรนน้ำ (Water Draen)"],
    "COMP-02": ["เช็คแรงดัน (Pressure) ต้องไม่ต่ำกว่า 7 bar", "ตรวจสอบระดับน้ำมันไฮดรอลิก ต้องไม่ต่ำกว่าระดับต่ำสุด", "เช็คอุณหภูมิความร้อนต้องไม่เกิน 80 องศา", "เช็คการรั่วซีมของระบบน้ำมัน", "เช็คระบบเดรนน้ำ (Water Draen)"],
    "GRINDING-01": ["การ Worm spindle และ TABLE SLIDE", "เช็คระดับนำมันไฮดรอลิก และ การทำงานของ PUMP", "เช็คระดับของน้ำยา COOLANNT PUMP", "ตรวจสอบการทำงานของแม่เหล็ก", "ตรวจสอบการทำงานของ SLIDE X,Y", "ตรวจสอบสภาพความพร้อมโดยรวมของเครื่องจักร", "ตรวจสอบระดับน้ำมันของ PUMPน้ำมันหล่อลื่น", "ตรวจสอบการทำงานของไฟฟ้าและแสงสว่าง", "ตรวจสอบการทำงานของตัวดูดอากศ"],
    "GRINDING-02": ["การ Worm spindle และ TABLE SLIDE", "เช็คระดับนำมันไฮดรอลิก และ การทำงานของ PUMP", "เช็คระดับของน้ำยา COOLANNT PUMP", "ตรวจสอบการทำงานของแม่เหล็ก", "ตรวจสอบการทำงานของ SLIDE X,Y", "ตรวจสอบสภาพความพร้อมโดยรวมของเครื่องจักร", "ตรวจสอบระดับน้ำมันของ PUMPน้ำมันหล่อลื่น", "ตรวจสอบการทำงานของไฟฟ้าและแสงสว่าง", "ตรวจสอบการทำงานของตัวดูดอากศ"],
    "CUTTER GRINDING-01": ["การ WORM UP แกน Y พร้อมใช้งาน", "การ WORM UP แกน Z พร้อมใช้งาน", "ตรวจสอบการทำงานของไฟฟ้าและแสงสว่าง", "ตรวจสอบการทำงานของมอเตอร์ มีการหมุนปกติ", "ตรวจสอบการจับหัวคอเรต"],
    "MILLING": [
        "Worm Spindle ก่อนเริมงาน ตรวจสอบความ ผิดปกติของชุด  Back gauge  และ Motor", 
        "เช็ค Auto  Up-Down back gauge  และ Manual ( ความคร่องตัวในการเคลื่อนที่ของ Spindle )", 
        "ตรวจสอบการ SLIDE  ของแกน X", "ตรวจสอบการ SLIDE  ของแกน Y", "ตรวจสอบการ SLIDE  ของแกน Z", 
        "ระดับน้ำมันไฮดรอลิค ตรวจสอบน้ำมันในปั้มน้ำมันหล่อ ลื่นแกน  X,Y,Z", 
        "ตรวจน้ำมันหล่อลื่นเย็น ตรวจสอบการทำงานของปั้ม COOLANT และสภาพของน้ำ  COOLANT", 
        "ตรวจสอบหน้าจอ  DIGITAL READ OUT และการทำ งานของ LINEAR SCALE", "หยอดน้ำมันหล่อลื่นทุกวันจันทร์", 
        "ตรวจสอบการทำงานของไฟฟ้าแสงสว่างของเครื่อง", "ตรวจสอบสภาพความพร้อมโดยรวมของเครื่องจักร  และอุกรณ์เสริมต่าง ๆ"
    ],
    "LATHE": [
        "Spindle ก่อนเริ่มงาน 15 นาที", "เช็คระดับน้ำมันเครื่องในห้องเกียร์", "เช็คระบบเฟื่องทดลองเปลี่ยนรอบตวามเร็วต่าง ๆ",
        "เช็ค AUTO แกน X,Y", "เช็คระดับน้ำมันหล่อลื่นใน PUMP", "เช็คน้ำมันหล่อเย็นและการทำงานของปั้ม",
        "ตรวจสอบหน้าจอ  DIGITAL  READ OUT และการ ทำงานของ  LINEAR  SCALE", "ตรวจเช็คสภาพและความตึงของสายพาน",
        "ตรวจสอบการทำงานของไฟฟ้าแสงสว่าง", "อัดจาระบีตามหัวอัดจาระบีทุก ๆ จุด", "ตรวจสอบความพร้อมสภาพโดยรวมของเครื่อง"
    ],
    "CUTTING": ["การ Worm spindle ก่อนเริ่มงาน เพื่อตรวจ ความผิดปกติของชุด Back gauge และ Motor", "เช็ค Auto Up-Down back gauge และ Manual ( ความคล่องตัวในการเคลื่อนที่ )", "ระดับน้ำมันไฮดรอลิค ตรวจสอบระดับในปั้มน้ำมัน หล่อลืนแกน  Back gauge", "ตรวจเช็ค  Switch  เปิด-ปิด", "ตรวจสอบ Digital  read out และการทำงาน ของ Linear  scale", "อัดจาระบีตามจุดที่อัดจาระบีทุกๆจุด", "ตรวจสอบใบมีด  บนและล่าง", "ตรวจสอบความพร้อมสภาพโดยรวมของเครื่อง จักรและอุปกรณ์เสริมต่าง ๆ"],
    "BENDING": [
        "การ Worm spindle ก่อนเริมงาน เพื่อตรวจสอบความ ผิดปกติของชุด  Back gauge  และ Motor",
        "เช็ค Auto  Up-Down back gauge  และ Manual ( ความคล่องตัวในการเคลื่อนที่ของ Spindle )",
        "ระดับน้ำมันไฮดรอลิค ตรวจสอบระดับน้ำมันในปั้ม น้ำมันหล่อลื่นแกน  Back gauge", "ตรวจเช็ค  Switch  เปิด-ปิด",
        "ตรวจสอบหน้าจอ Digital read out  และการทำงาน ของ Linear  scale", "อัดจาระบีตามจุดหัวอัดจาระบีทุก ๆจุด 1ครั้งตต่อเดือน",
        "ตรวจสอบการทำงานของไฟฟ้าแสงสว่างของเครื่อง", "ตรวจสอบฟันพับของร่อง  V",
        "ตรวจสอบความพร้อมและสภาพโดยรวมของเครื่องจักรและอุกรณ์เสริมต่าง ๆ"
    ],
    "MIG CO2": ["ตรวจสภาพความพร้อมโดยรวมของเครื่อง", "เช็ค BREAKER เพื่อเช็คระบบไฟฟ้า ตามตำแหน่งไฟ โชว์ และสวิชท์ต่าง ๆ", "ตรวจสภาพความพร้อมของมาตราวัดแรงดัน ของก๊าซ CO2 และปรับตั้งอย่างถูกต้อง", "ตรวจจุดต่อของก๊าซ CO2 รั่วหรือไม่", "ตรวจสภาพความพร้อมของสายไฟ สายก๊าซ  CO2 ว่ารั่วหรือไม่", "ตรวจสภาพความพร้อมของสายกราวด์", "ทำความสะอาดหัวเชื่อมก่อนใช้งาน"],
    "ARGON": [
        "ตรวจสภาพความพรัอมโดยรวมของเครื่อง", "เช็ค  BREAKER  เพื่อเช็คระบบไฟฟ้า ตามตำแหน่งไฟ โชว์  และ SWITCH  ต่าง ๆ", 
        "ตรวจสภาพความพร้อมของมาตราวัดแรงดันของมาตรา วัดแรงดันของก๊าช  ARGON  และปรับตั้งอย่างถูกวิธี", "ตรวจุดต่อของสายก๊าช  ARGON  ก่อนว่ารั่วหรือไม่", 
        "ตรวจสภาพความพร้อมของสายกราว์", "ตรวจสภาพความพร้อมของสายไฟฟ้าสายก๊าช  ARGON และชุดหัวเชื่อม", 
        "ตรวจสภาพความพร้อมของ  SWITCH  หัวเชื่อม", "ทำความสะดาดชุดหัวเชื่อมก่อนใช้งาน"
    ],
    "WELDING_ALUMINUM": [
        "ตรวจสภาพความพรัอมโดยรวมของเครื่อง", "เช็ค  BREAKER  เพื่อเช็คระบบไฟฟ้า ตามตำแหน่งไฟ โชว์  และ SWITCH  ต่าง ๆ",
        "ตรวจสภาพความพร้อมของสายกราวด์ให้อยู่ในสภาพ ความพร้อมอยู่เสมอ", "ตรวจจุดต่อของสายท่อแก๊สว่ารั่วหรือไม่",
        "เช็คระดับน้ำหล่อเย็นให้อยู่ในระดับพร้อมใช้งาน", "เช็คระดับแรงดันในถังแก๊สให้พร้อมใช้งาน",
        "ทำความสะอาดชุดหัวเชื่อมก่อนใช้งานอย่างสม่ำเสมอ"
    ],
    "BAND SAW": ["เช็ค Auto Up-Down Back Gauge และ Manual (ความคล่องตัวในการเคลื่อนที่ of Spindle)", "เช็คระดับน้ำมันไฮดรอลิค", "ตรวจน้ำมันหล่อลื่นเย็น ตรวจสอบการทำงานของปั๊ม COOLANT และสภาพของน้ำ COOLANT", "ตรวจสอบ Switch (สวิตซ์) หน้า BOX CONTROL", "ตรวจสอบระดับน้ำมันหล่อลื่นในห้องเกียร์"],
    "FORKLIFT": [
        "ตรวจเช็คระบบน้ำหม้อน้ำให้อยู่ในระดับ Hight", "ตรวจเช็คน้ำมันเครื่องยนต์ต้องอยู่ไม่เกินขีดที่3ของตัวเช็ค", "ตรวจเช็คไส้กรองและเป่าลมทำความสะอาด",
        "ตรวจเช็คการรั่วซึมofน้ำมันไฮดรอริก", "ตรวจเช็คระบบเบรคและน้ำมันเบรค", "ตรวจเช็คไฟส่องสว่างและไฟเลี้ยว",
        "ตรวจเช็คสัญญานแตร"
    ]
}

PHOTO_RULES = {
    "CNC": [2, 3, 4, 5, 8, 13], "Crane no.1": [3, 4], "Crane no.2": [3, 4], "QC-01": [4],
    "QC-02": [2, 4], "QC-03": [2, 4], "QC-04": [2, 4], "QC-05": [2, 4], "QC-06": [2, 4],
    "QC-07": [2, 4], "QC-08": [2, 4], "QC-09": [2, 4], "QC-10": [2], "QC-11": [2], "QC-12": [2],
    "QC-13": [2], "QC-14": [2], "QC-15": [6], "QC-16": [3], "QC-17": [2], "QC-18": [3], "QC-19": [3],
    "QC-20": [3], "QC-21": [3], "COMP-01": [1, 2, 3], "COMP-02": [1, 2, 3], "GRINDING-01": [2, 4, 7], "GRINDING-02": [4, 7],
    "CUTTER GRINDING-01": [], "MILLING": [6, 7], "LATHE": [2, 5, 6], "CUTTING": [3, 5, 7], "BENDING": [3, 5, 6], "MIG CO2": [3, 4, 5],
    "ARGON": [3, 4, 6], "WELDING_ALUMINUM": [5, 6], "BAND SAW": [2, 3, 5], "FORKLIFT": [1, 2, 5]
}

def get_coordinates_by_machine(m_id, m_type):
    u_id = str(m_id).upper()
    if "CUTTER" in u_id or m_type == "CUTTER GRINDING-01": return 13, 15, "B18"
    if "QC-01" in u_id: return 10, 12, "B15"
    if any(k in u_id for k in ["QC-02", "QC-03", "QC-04", "QC-05", "QC-06", "QC-07", "QC-08", "QC-09", "QC-13", "QC-14", "QC-16", "QC-17", "QC-18", "QC-19", "QC-20", "QC-21"]): return 11, 13, "B16"
    if any(k in u_id for k in ["QC-10", "QC-11", "QC-12"]): return 11, 13, "B15"
    if "QC-15" in u_id: return 12, 14, "B17"
    if "ARGON-02" in u_id or "ARGON-01" in u_id: return 14, 16, "B19"
    if m_type == "FORKLIFT" or "FORKLIFT" in u_id: return 13, 15, "B18"
    if m_type == "CNC" or "CNC" in u_id: return 22, 24, "B28"
    if "CRANE" in u_id: return 14, 16, "B19"
    if "GRINDING" in m_type or "GRINDING" in u_id: return 16, 18, "B21"
    if m_type == "MILLING" or "MILLING" in u_id: return 17, 19, "B22" 
    if m_type == "LATHE" or "LATHE" in u_id: return 17, 19, "B22"
    if m_type == "CUTTING" or "CUTTING" in u_id: return 13, 15, "B18"
    if m_type == "BENDING" or "BENDING" in u_id: return 15, 17, "B20" 
    if m_type == "WELDING_ALUMINUM" or "WELDING_ALUMINUM" in u_id: return 13, 15, "B18"
    if m_type == "MIG CO2" or "MIG" in u_id: return 13, 15, "B18"
    if m_type == "BAND SAW" or "BAND" in u_id: return 11, 13, "B16"
    return 11, 13, "B16"

def get_unmerged_cell(ws, coordinate_str):
    cell = ws[coordinate_str]
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell

# --- 2. 🛡️ ฐานข้อมูลกระจกเงาคู่ขนานรักษาประวัติถาวร ---
def save_log_to_mirror_db(machine_id, day_num, year_month, tech_name, checklist_item, item_no, status, note, role="tech"):
    local_cloud_backup = os.path.join(BASE_FOLDER, "gsheet_cloud_mirror.csv")
    new_data = {
        "Timestamp": [datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        "Machine_ID": [machine_id],
        "Day_Num": [int(day_num)],
        "Year_Month": [year_month],
        "Tech_Name": [tech_name],
        "Item_No": [int(item_no)],
        "Checklist_Item": [checklist_item],
        "Status": [status],
        "Note": [note],
        "Role": [role]
    }
    df_new = pd.DataFrame(new_data)
    if not os.path.exists(local_cloud_backup):
        df_new.to_csv(local_cloud_backup, index=False, encoding="utf-8-sig")
    else:
        df_new.to_csv(local_cloud_backup, mode='a', header=False, index=False, encoding="utf-8-sig")

def fetch_logs_from_mirror_db(machine_id, year_month):
    local_cloud_backup = os.path.join(BASE_FOLDER, "gsheet_cloud_mirror.csv")
    if not os.path.exists(local_cloud_backup):
        return pd.DataFrame()
    try:
        df = pd.read_csv(local_cloud_backup, encoding="utf-8-sig")
        df_filtered = df[(df["Machine_ID"] == machine_id) & (df["Year_Month"] == year_month)]
        return df_filtered
    except:
        return pd.DataFrame()

def apply_mirror_history_to_excel(machine_id, year_month, m_type):
    """🎯 แก้ไขจุดหัวใจสำคัญ: กวาดข้อมูลกระจกเงาเขียนลง Excel แบบรวมวันเดียว ไม่เบิ้ลซ้ำ!"""
    target_excel_path = os.path.join(BASE_FOLDER, f"FM-MN-07_{machine_id}.xlsx")
    if not os.path.isfile(target_excel_path):
        return
    df_logs = fetch_logs_from_mirror_db(machine_id, year_month)
    if df_logs.empty:
        return
    try:
        wb = openpyxl.load_workbook(target_excel_path, data_only=False)
        ws = wb.active
        t_row, boss_row, n_cell = get_coordinates_by_machine(machine_id, m_type)
        
        # 1. เคลียร์ช่อง Note B ก่อนรวมประวัติใหม่ทั้งหมด ป้องกันการเขียนเบิ้ลสะสม
        note_cell = get_unmerged_cell(ws, n_cell)
        note_cell.value = ""
        
        df_logs = df_logs.sort_values(by="Timestamp")
        notes_by_day = {}
        
        for _, row in df_logs.iterrows():
            day_val = int(row["Day_Num"])
            col_letter = get_column_letter(2 + day_val)
            
            if row["Role"] == "tech":
                status_val = row["Status"]
                item_idx = int(row["Item_No"])
                cell_coordinate = f"{col_letter}{5 + item_idx}"
                current_cell = get_unmerged_cell(ws, cell_coordinate)
                
                if status_val == "ใช้งานได้ปกติ": current_cell.value = "/"
                elif status_val == "ทำการแก้ไขใช้งานได้ปกติ": current_cell.value = "⨂"
                elif status_val == "ใช้งานไม่ได้ต้องแก้ไข": current_cell.value = "X"
                elif status_val == "ไม่ได้ทำงาน": current_cell.value = "-"
                current_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                tech_cell = get_unmerged_cell(ws, f"{col_letter}{t_row}")
                tech_cell.value = row["Tech_Name"]
                tech_cell.alignment = Alignment(text_rotation=90, horizontal='center', vertical='center')
                
                # เก็บข้อความแยกตามวัน ไม่ให้เบิ้ลซ้ำ
                if str(row["Note"]).strip() and str(row["Note"]) != "nan":
                    if day_val not in notes_by_day:
                        notes_by_day[day_val] = []
                    if str(row["Note"]).strip() not in notes_by_day[day_val]:
                        notes_by_day[day_val].append(str(row["Note"]).strip())

            elif row["Role"] == "boss":
                boss_cell = get_unmerged_cell(ws, f"{col_letter}{boss_row}")
                boss_cell.value = row["Tech_Name"]
                boss_cell.alignment = Alignment(text_rotation=90, horizontal="center", vertical="center")

        # 2. นำข้อความที่รวบรวมได้ พ่นลงช่อง Note B บรรทัดเดียวจบ สวยงาม
        final_notes_list = []
        for d_key in sorted(notes_by_day.keys()):
            day_txt = f"[วันที่ {d_key}]: " + ", ".join(notes_by_day[d_key])
            final_notes_list.append(day_txt)
            
        if final_notes_list:
            note_cell.value = ",  ".join(final_notes_list)
            note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        wb.save(target_excel_path)
    except:
        pass

# --- PHOTO & ZIP FUNCTIONS ---
def send_line_alert(msg_text):
    url = 'https://api.line.me/v2/bot/message/push'
    headers = {'Content-Type': 'application/json', 'Authorization': f'Bearer {LINE_ACCESS_TOKEN}'}
    payload = {"to": LINE_TARGET_ID, "messages": [{"type": "text", "text": msg_text}]}
    try: requests.post(url, headers=headers, data=json.dumps(payload))
    except Exception as e: print(f"ส่งไลน์ไม่สำเร็จ: {e}")

def save_uploaded_photos_list(machine_id, day_num, item_index, files_list, current_date_obj=None):
    saved_paths = []
    if files_list:
        if current_date_obj is None:
            current_date_obj = datetime.date.today()
        current_year_month = current_date_obj.strftime("%Y_%B")
        
        folder_path = os.path.join(BASE_FOLDER, "maintenance_photos", str(machine_id), current_year_month, f"Day_{day_num}")
        if not os.path.exists(folder_path): os.makedirs(folder_path, exist_ok=True)
        
        for idx, uploaded_file in enumerate(files_list, 1):
            file_extension = os.path.splitext(uploaded_file.name)[1]
            file_name = f"photo_item_{item_index}_{idx}{file_extension}"
            full_path = os.path.join(folder_path, file_name)
            with open(full_path, "wb") as f: f.write(uploaded_file.getbuffer())
            saved_paths.append(full_path)
    return saved_paths

def zip_single_machine_photos(machine_id, target_date_obj, target_day=None):
    current_year_month = target_date_obj.strftime("%Y_%B")
    if target_day:
        source_dir = os.path.join(BASE_FOLDER, "maintenance_photos", str(machine_id), current_year_month, f"Day_{target_day}")
    else:
        source_dir = os.path.join(BASE_FOLDER, "maintenance_photos", str(machine_id), current_year_month)
        
    if not os.path.exists(source_dir):
        return None
        
    zip_buffer = BytesIO()
    try:
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(source_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    relative_path = os.path.relpath(file_path, source_dir)
                    zip_file.write(file_path, relative_path)
        zip_buffer.seek(0)
        return zip_buffer
    except:
        return None

def zip_all_factory_photos_by_filter(filter_type="ทั้งโรงงาน", target_date_obj=None):
    photos_root_dir = os.path.join(BASE_FOLDER, "maintenance_photos")
    if not os.path.exists(photos_root_dir):
        return None
    if target_date_obj is None:
        target_date_obj = datetime.date.today()
    current_year_month = target_date_obj.strftime("%Y_%B")
        
    zip_buffer = BytesIO()
    try:
        has_file = False
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for machine_code in MACHINES.keys():
                match = False
                if filter_type == "ทั้งโรงงาน": match = True
                elif filter_type == "CNC" and "CNC" in machine_code: match = True
                elif filter_type == "GRINDING" and "GRINDING" in machine_code.upper(): match = True
                elif filter_type == "CRANE" and "CRANE" in machine_code.upper(): match = True
                elif filter_type == "COMPRESSOR" and "COMP-" in machine_code.upper(): match = True
                elif filter_type == "QC" and "QC-" in machine_code.upper(): match = True
                elif filter_type == "MILLING" and "MILLING" in machine_code: match = True
                elif filter_type == "MIG CO2" and "MIG" in machine_code: match = True
                elif filter_type == "ARGON" and "ARGON" in machine_code: match = True
                elif filter_type == "เครื่องจักรอื่น ๆ (พับ/ตัด/กลึง/โฟคลิฟ)" and any(k in machine_code for k in ["BENDING", "CUTTING", "LATHE", "FORKLIFT", "WELDING_ALUMINUM", "SAW"]): match = True
                
                if match:
                    machine_dir = os.path.join(photos_root_dir, machine_code, current_year_month)
                    if os.path.exists(machine_dir):
                        for root, dirs, files in os.walk(machine_dir):
                            for file in files:
                                file_path = os.path.join(root, file)
                                arc_name = os.path.join(machine_code, os.path.relpath(file_path, machine_dir))
                                zip_file.write(file_path, arc_name)
                                has_file = True
        if not has_file:
            return None
        zip_buffer.seek(0)
        return zip_buffer
    except:
        return None

def update_iso_excel_by_tech(machine_id, day_num, results_dict, tech_name, m_type):
    target_excel_path = os.path.join(BASE_FOLDER, f"FM-MN-07_{machine_id}.xlsx")
    if not os.path.isfile(target_excel_path): return False, f"ไม่พบไฟล์แบบฟอร์ม `{target_excel_path}` บนระบบคลาวด์"
    try:
        wb = openpyxl.load_workbook(target_excel_path, data_only=False)
        ws = wb.active
        
        t_row, boss_row, n_cell = get_coordinates_by_machine(machine_id, m_type)
        check_col_1 = get_column_letter(3) 
        first_cell_of_month = get_unmerged_cell(ws, f"{check_col_1}{t_row}")
        
        # ล้างตารางเริ่มเดือนใหม่
        if day_num == 1 and (first_cell_of_month.value is None or first_cell_of_month.value == ""):
            backup_folder = os.path.join(BASE_FOLDER, "maintenance_backups")
            if not os.path.exists(backup_folder): os.makedirs(backup_folder, exist_ok=True)
            today = datetime.date.today()
            first_day_of_current_month = today.replace(day=1)
            last_day_of_last_month = first_day_of_current_month - datetime.timedelta(days=1)
            last_month_str = last_day_of_last_month.strftime("%B_%Y")
            
            backup_file_name = f"Backup_{last_month_str}_FM-MN-07_{machine_id}.xlsx"
            backup_excel_path = os.path.join(backup_folder, backup_file_name)
            
            if not os.path.exists(backup_excel_path):
                shutil.copy2(target_excel_path, backup_excel_path)
                try: send_line_alert(f"📦 [Auto-Backup]: สำรองไฟล์เครื่อง {machine_id} แล้ว")
                except: pass

            checklist_items = CHECKLISTS[m_type]
            for d in range(1, 32):
                c_letter = get_column_letter(2 + d)
                for row_idx in range(6, 6 + len(checklist_items)):
                    ws.cell(row=row_idx, column=2 + d, value="")
                get_unmerged_cell(ws, f"{c_letter}{t_row}").value = ""
                get_unmerged_cell(ws, f"{c_letter}{boss_row}").value = ""
            
            note_cell = get_unmerged_cell(ws, n_cell)
            note_cell.value = ""
            note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        wb.save(target_excel_path)
        return True, ""
    except Exception as e:
        return False, str(e)

def approve_excel_by_boss(machine_id, day_num, boss_name, m_type):
    target_excel_path = os.path.join(BASE_FOLDER, f"FM-MN-07_{machine_id}.xlsx")
    if not os.path.isfile(target_excel_path): return False
    try:
        wb = openpyxl.load_workbook(target_excel_path, data_only=False)
        ws = wb.active
        col_letter = get_column_letter(2 + day_num)
        _, boss_row, _ = get_coordinates_by_machine(machine_id, m_type)
        
        boss_cell = get_unmerged_cell(ws, f"{col_letter}{boss_row}")
        boss_cell.value = boss_name
        boss_cell.alignment = Alignment(text_rotation=90, horizontal="center", vertical="center")
        wb.save(target_excel_path)
        return True
    except Exception as e: print(f"Boss approve error: {e}"); return False

def get_current_excel_note(machine_id, m_type):
    target_excel_path = os.path.join(BASE_FOLDER, f"FM-MN-07_{machine_id}.xlsx")
    if not os.path.isfile(target_excel_path): return ""
    try:
        wb = openpyxl.load_workbook(target_excel_path, data_only=False)
        ws = wb.active
        _, _, n_cell = get_coordinates_by_machine(machine_id, m_type)
        note_cell = get_unmerged_cell(ws, n_cell)
        val = note_cell.value
        return val if val else ""
    except: return ""

# --- 3. UI NAVIGATION SIDEBAR & QUERY PARAMETERS ---
st.set_page_config(page_title="Smart Factory PM SYSTEM", page_icon="🔧", layout="wide")

query_params = st.query_params
raw_role = query_params.get("role", "tech")
is_boss_link = str(raw_role).strip().lower() == "boss"

if is_boss_link:
    user_role = "🔐 หัวหน้างาน/ผู้ตรวจสอบ"
else:
    st.sidebar.title("🏢 เมนูควบคุมโรงงานรวม")
    user_role = st.sidebar.radio("เลือกสิทธิ์การเข้าใช้งานด้านล่าง:", ["🔧 ช่างเทคนิค (ส่งฟอร์ม)", "🔐 หัวหน้างาน/ผู้ตรวจสอบ"])

now = datetime.datetime.now()
current_time_str = now.strftime("%Y-%m-%d %H:%M:%S")

raw_machine_id = query_params.get("id", "CNC3X-01")
if isinstance(raw_machine_id, list): machine_id = str(raw_machine_id[0]).strip()
else: machine_id = str(raw_machine_id).strip()
machine_id = machine_id.replace("%20", " ")

if "CUTTER" in machine_id.upper(): m_type_selected = "CUTTER GRINDING-01"
elif "CRANE NO.1" in machine_id.upper() or "CRANE no.1" in machine_id: m_type_selected = "Crane no.1"
elif "CRANE NO.2" in machine_id.upper() or "CRANE no.2" in machine_id: m_type_selected = "Crane no.2"
elif "QC-01" in machine_id.upper() or "QC-01" in machine_id: m_type_selected = "QC-01"
elif "QC-02" in machine_id.upper() or "QC-02" in machine_id: m_type_selected = "QC-02"
elif "QC-03" in machine_id.upper() or "QC-03" in machine_id: m_type_selected = "QC-03"
elif "QC-04" in machine_id.upper() or "QC-04" in machine_id: m_type_selected = "QC-04"
elif "QC-05" in machine_id.upper() or "QC-05" in machine_id: m_type_selected = "QC-05"
elif "QC-06" in machine_id.upper() or "QC-06" in machine_id: m_type_selected = "QC-06"
elif "QC-07" in machine_id.upper() or "QC-07" in machine_id: m_type_selected = "QC-07"
elif "QC-08" in machine_id.upper() or "QC-08" in machine_id: m_type_selected = "QC-08"
elif "QC-09" in machine_id.upper() or "QC-09" in machine_id: m_type_selected = "QC-09"
elif "QC-10" in machine_id.upper() or "QC-10" in machine_id: m_type_selected = "QC-10"
elif "QC-11" in machine_id.upper() or "QC-11" in machine_id: m_type_selected = "QC-11"
elif "QC-12" in machine_id.upper() or "QC-12" in machine_id: m_type_selected = "QC-12"
elif "QC-13" in machine_id.upper() or "QC-13" in machine_id: m_type_selected = "QC-13"
elif "QC-14" in machine_id.upper() or "QC-14" in machine_id: m_type_selected = "QC-14"
elif "QC-15" in machine_id.upper() or "QC-15" in machine_id: m_type_selected = "QC-15"
elif "QC-16" in machine_id.upper() or "QC-16" in machine_id: m_type_selected = "QC-16"
elif "QC-17" in machine_id.upper() or "QC-17" in machine_id: m_type_selected = "QC-17"
elif "QC-18" in machine_id.upper() or "QC-18" in machine_id: m_type_selected = "QC-18"
elif "QC-19" in machine_id.upper() or "QC-19" in machine_id: m_type_selected = "QC-19"
elif "QC-20" in machine_id.upper() or "QC-20" in machine_id: m_type_selected = "QC-20" 
elif "QC-21" in machine_id.upper() or "QC-21" in machine_id: m_type_selected = "QC-21"
elif "COMP-01" in machine_id.upper(): m_type_selected = "COMP-01"
elif "COMP-02" in machine_id.upper(): m_type_selected = "COMP-02"
elif "GRINDING-01" in machine_id.upper(): m_type_selected = "GRINDING-01"
elif "GRINDING-02" in machine_id.upper(): m_type_selected = "GRINDING-02"
elif "MILLING" in machine_id.upper(): m_type_selected = "MILLING"
elif "LATHE" in machine_id.upper(): m_type_selected = "LATHE"
elif "CUTTING" in machine_id.upper(): m_type_selected = "CUTTING"
elif "BENDING" in machine_id.upper(): m_type_selected = "BENDING" 
elif "MIG" in machine_id.upper(): m_type_selected = "MIG CO2"
elif "ARGON" in machine_id.upper(): m_type_selected = "ARGON"
elif "WELDING_ALUMINUM" in machine_id.upper(): m_type_selected = "WELDING_ALUMINUM"
elif "BAND" in machine_id.upper(): m_type_selected = "BAND SAW"
elif "FORKLIFT" in machine_id.upper(): m_type_selected = "FORKLIFT"
else: m_type_selected = "CNC"

# ==========================================
# 🔧 [โหมดที่ 1: ฝั่งช่างเทคนิคส่งฟอร์มประจำวัน]
# ==========================================
if user_role == "🔧 ช่างเทคนิค (ส่งฟอร์ม)":
    st.image("Logo_Pes.png", width=240)
    st.caption("PHOLLAWAT ENGINEERING SUPPLY CO., LTD.")
    st.title(f"📋 ใบตรวจสอบเครื่อง {machine_id} ประจำวัน")
    st.info("📄 มาตรฐานระบบคุณภาพโรงงาน: **FM-MN-07 Rev.00**")

    if machine_id in MACHINES: st.success(f"⚙️ คุณกำลังตรวจเครื่อง: **{machine_id} ({MACHINES[machine_id]})**")
    else: st.error(f"⚠️ ไม่พบรหัสเครื่อง '{machine_id}' ในทะเบียนกลาง")
    st.divider()

    report_date = st.date_input("📆 เลือกวันที่ตรวจสอบงานฟอร์ม:", value=datetime.date.today())
    current_day = report_date.day
    year_month_key = report_date.strftime("%Y_%B")

    with st.form("pm_form"):
        tech_name = st.text_input("👤 ชื่อช่างผู้ตรวจเช็ค (ผู้รับผิดชอบ)", placeholder="ระบุชื่อ-นามสกุลของคุณ")
        results, uploaded_photos = {}, {}
        current_checklist = CHECKLISTS[m_type_selected]
        required_photo_indexes = PHOTO_RULES[m_type_selected]
        
        for i, item in enumerate(current_checklist, 1):
            st.write(f"**{i}. {item}**")
            status = st.radio(f"ผลการตรวจข้อ {i}", ["ใช้งานได้ปกติ", "ทำการแก้ไขใช้งานได้ปกติ", "ใช้งานไม่ได้ต้องแก้ไข", "ไม่ได้ทำงาน"], horizontal=True, key=f"check_{i}", label_visibility="collapsed", index=None)
            if i in required_photo_indexes:
                st.write("📷 *หัวข้อบังคับถ่ายรูปหลักฐานยืนยันหน้างานจริง (เลือกได้มากกว่า 1 รูป)*")
                uploaded_files = st.file_uploader(f"แนบรูปข้อ {i}", type=["jpg", "jpeg", "png"], key=f"photo_{i}", accept_multiple_files=True)
                uploaded_photos[i] = {"files": uploaded_files, "index": i}
            note = st.text_input(f"หมายเหตุ/อาการเสีย (ข้อ {i})", key=f"note_{i}", placeholder="ระบุรายละเอียดหากพบจุดพังหรือบันทึกงานซ่อมแก้ไข")
            results[item] = {"status": status, "note": note}
            st.divider()

        submitted = st.form_submit_button("💾 ส่งรายงานการตรวจเช็คประจำวัน (SUBMIT)")

    if submitted:
        if machine_id not in MACHINES: st.error("❌ รหัสเครื่องจักรไม่ถูกต้อง")
        elif not tech_name: st.error("❌ กรุณาระบุชื่อผู้ตรวจสอบก่อนส่งรายงานครับ!")
        elif any(results[item]["status"] is None for item in current_checklist): st.error("❌ ปฏิเสธการบันทึก! ช่างยังเลือกผลการตรวจสอบไม่ครบทุกหัวข้อ")
        elif any((uploaded_photos[idx]["files"] is None or len(uploaded_photos[idx]["files"]) == 0) for idx in required_photo_indexes): st.error(f"❌ ปฏิเสธการบันทึกฟอร์ม! กรุณาถ่ายภาพหลักฐานประจำข้อ {required_photo_indexes} ให้ครบถ้วนก่อนกดส่งครับ")
        else:
            photo_logs = []
            for idx in required_photo_indexes:
                saved_paths = save_uploaded_photos_list(machine_id, current_day, idx, uploaded_photos[idx]["files"], current_date_obj=report_date)
                if saved_paths: 
                    photo_logs.append(f"📸 แนบรูปหลักฐานข้อ {idx} สำเร็จ ({len(saved_paths)} รูป)")
            
            # บันทึกประวัติลงฐานข้อมูลเงา
            for i, item in enumerate(current_checklist, 1):
                save_log_to_mirror_db(machine_id, current_day, year_month_key, tech_name, item, i, results[item]["status"], results[item]["note"], "tech")

            fails, fixed_items = [], []
            for i, item in enumerate(current_checklist, 1):
                status_val = results[item]["status"]
                note_val = results[item]["note"]
                if status_val == "ใช้งานไม่ได้ต้องแก้ไข": fails.append(f"- ข้อ {i}. {item}" + (f" ({note_val})" if note_val else ""))
                elif status_val == "ทำการแก้ไขใช้งานได้ปกติ": fixed_items.append(f"- ข้อ {i}. {item}" + (f" ({note_val})" if note_val else ""))
            
            update_iso_excel_by_tech(machine_id, current_day, results, tech_name, m_type_selected)
            apply_mirror_history_to_excel(machine_id, year_month_key, m_type_selected)
            
            boss_review_url = f"https://pes-maintenance.streamlit.app/?role=boss&id={machine_id}"
            audit_tag = f"\n\n📂 [คลิกเปิดตรวจรายงานและดูภาพหลักฐานคลาวด์]:\n👉 {boss_review_url}"
            
            if fails:
                summary_msg = f"\n🚨 [แจ้งซ่อมด่วนจากใบตรวจเช็ค ISO]\n🔧 เครื่อง: {MACHINES[machine_id]}\n📅 วันที่: {current_time_str}\n👤 ผู้ตรวจ: {tech_name}\n\n❌ รายการที่ไม่ผ่านมาตรฐาน:\n" + "\n".join(fails)
                if fixed_items: summary_msg += "\n\n🛠️ รายการที่ช่างแก้ไขเสร็จทันที:\n" + "\n".join(fixed_items)
                send_line_alert(summary_msg + audit_tag)
                st.warning("พบจุดบกพร่อง! ส่งการแจ้งเตือนเตือนเข้าไลน์กลุ่มช่างแล้ว")
            else:
                ok_msg = f"\n🎉 [รายงานเครื่องจักรปกติ - ISO]\n🔧 เครื่อง: {MACHINES[machine_id]}\n📅 วันที่: {current_time_str}\n✅ ผลการตรวจสอบ: ปกติทุกหัวข้อ\n👤 ผู้ตรวจสอบ: {tech_name}"
                if fixed_items: ok_msg += "\n\n🛠️ รายการที่ช่างแก้ไขหน้างานสำเร็จ (ลงตาราง ⨂):\n" + "\n".join(fixed_items)
                send_line_alert(ok_msg + audit_tag)
            st.success(f"🎉 บันทึกรายงานเครื่อง {machine_id} สำเร็จ! ข้อมูลอัปเดตเรียบร้อยแล้ว")

# ==========================================
# 🔐 [โหมดที่ 2: ฝั่งหัวหน้างาน ล็อกอินตรวจสอบและกดอนุมัติฟอร์ม]
# ==========================================
else:
    st.image("Logo_Pes.png", width=240)
    st.caption("PHOLLAWAT ENGINEERING SUPPLY CO., LTD.")
    st.title("🔐 หน้าต่างควบคุมระบบตรวจสอบคุณภาพ (สำหรับหัวหน้างาน)")
    
    selected_date = st.date_input("📆 เลือกวันที่ต้องการตรวจสอบเอกสารและดูรูปภาพย้อนหลัง:", value=datetime.date.today())
    target_day_check = selected_date.day
    year_month_key = selected_date.strftime("%Y_%B")
    
    st.subheader(f"📅 ประจำวันที่เลือก: {selected_date.strftime('%d/%m/%Y')} (คอลัมน์ Excel ช่องวันที่ {target_day_check})")
    
    is_supervisor = False
    is_bigboss = False
    
    password_input = st.text_input("🔑 กรุณากรอกรหัสผ่านหลักเพื่อเข้าสู่ระบบบอร์ดควบคุมหลัก:", type="password")
    
    if password_input != "":
        is_supervisor = (password_input == BOSS_PASSWORD)
        is_bigboss = (password_input == BIGBOSS_PASSWORD)

    if is_supervisor or is_bigboss:
        if is_bigboss:
            st.success("👑 [สิทธิ์ผู้บริหารสูงสุด]: ล็อกอินผ่านรหัสแอนมินหลักเรียบร้อย")
            boss_name = st.text_input("👤 ชื่อผู้ตรวจสอบ/บิ๊กบอส:", value="พลวัฒน์ (Big Boss)")
        else:
            st.success("🔓 ยืนยันสิทธิ์: เข้าสู่ระบบตรวจสอบและบันทึกประจำวันได้")
            boss_name = st.text_input("👤 ชื่อผู้ตรวจสอบ/หัวหน้างาน:", value="พลวัฒน์")
            
        st.divider()
        st.write("### 📊 บอร์ดควบคุมการรายงานตรวจเช็ค ทั้งโรงงาน")
   
        def render_machine_card(m_id, m_name, m_type_flag):
            target_excel_path = os.path.join(BASE_FOLDER, f"FM-MN-07_{m_id}.xlsx")
            apply_mirror_history_to_excel(m_id, year_month_key, m_type_flag)
            
            st.info(f"⚙️ **{m_id}**\n{m_name}")
            
            if os.path.isfile(target_excel_path):
                if st.button(f"✅ อนุมัติฟอร์มของ {m_id}", key=f"btn_{m_id}"):
                    if approve_excel_by_boss(m_id, target_day_check, boss_name, m_type_flag):
                        save_log_to_mirror_db(m_id, target_day_check, year_month_key, boss_name, "BOSS APPROVAL", 0, "APPROVED", "", "boss")
                        st.toast(f"ลงนามดิจิทัลเครื่อง {m_id} สำเร็จ!", icon="🔥")
                        send_line_alert(f"🔒 [ISO Approved]: หัวหน้างาน ({boss_name}) ได้อนุมัติใบตรวจประจำวันที่ {target_day_check} ของเครื่อง {m_id} แล้ว")
                        st.success(f"✍️ เซ็นรับรองลงช่องผู้ตรวจสอบเครื่อง {m_id} สำเร็จ!")
            
            target_year_month_folder = selected_date.strftime("%Y_%B")
            img_dir = os.path.join(BASE_FOLDER, "maintenance_photos", str(m_id), target_year_month_folder, f"Day_{target_day_check}")
            
            if os.path.exists(img_dir):
                valid_photos = [os.path.join(img_dir, p) for p in os.listdir(img_dir) if p.lower().endswith(('.png', '.jpg', '.jpeg'))]
                if valid_photos:
                    with st.expander(f"📸 ตรวจรูปภาพหลักฐานวันที่ {target_day_check} ({len(valid_photos)} รูป)"):
                        for p_path in sorted(valid_photos):
                            st.image(p_path, caption=f"หลักฐาน: {os.path.basename(p_path)}", use_container_width=True)
            else:
                st.caption(f"ℹ️ วันที่ {target_day_check} ไม่มีรูปภาพหลักฐาน")

            current_notes = get_current_excel_note(m_id, m_type_flag)
            u_id = str(m_id).upper()
            if "CUTTER" in u_id or m_type_flag == "CUTTER GRINDING-01": note_label = "ช่อง B18"
            elif "ARGON-02" in u_id or "ARGON-01" in u_id or "CRANE" in u_id: note_label = "ช่อง B19"
            elif "WELDING_ALUMINUM" in u_id or "FORKLIFT" in u_id or "CUTTING" in u_id: note_label = "ช่อง B18"
            elif "CNC" in u_id: note_label = "ช่อง B28"
            elif "QC-01" in u_id or "QC-10" in u_id or "QC-11" in u_id or "QC-12" in u_id: note_label = "ช่อง B15"
            elif "QC-15" in u_id: note_label = "ช่อง B17"
            elif "GRINDING" in u_id or "GRINDING" in m_type_flag: note_label = "ช่อง B21"
            elif "MILLING" in u_id or "LATHE" in u_id: note_label = "ช่อง B22"
            elif "BENDING" in u_id: note_label = "ช่อง B20"
            else: note_label = "ช่อง B16"
            
            st.text_area(f"📝 รายการอาการเสียสะสม ({note_label}) [อ่านข้อมูลโหมดสัมปทานอัตโนมัติ]", value=current_notes, key=f"note_area_{m_id}", height=120, disabled=True)

            st.write("---")
            st.caption(f"📅 **เลือกดาวน์โหลดรูปภาพของ {m_id}:**")
            photo_date_input = st.date_input("เลือกวันที่ต้องการดึงรูปภาพ (.zip):", value=selected_date, key=f"photo_date_{m_id}")
            chosen_day = photo_date_input.day
            
            excel_col, zip_day_col, zip_month_col = st.columns(3)
            
            with excel_col:
                if os.path.isfile(target_excel_path):
                    with open(target_excel_path, "rb") as f:
                        st.download_button(label=f"📥 ดึง Excel {m_id}", data=f, file_name=f"FM-MN-07_{m_id}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_{m_id}")
                else:
                    st.button(f"❌ ไม่มีไฟล์ Excel", disabled=True, key=f"dl_disabled_{m_id}")
                    
            with zip_day_col:
                zip_day_data = zip_single_machine_photos(m_id, target_date_obj=photo_date_input, target_day=chosen_day)
                if zip_day_data:
                    st.download_button(label=f"📸 โหลดรูปวันที่ {chosen_day}", data=zip_day_data, file_name=f"Photos_{m_id}_Day_{chosen_day}.zip", mime="application/zip", key=f"zip_day_btn_{m_id}")
                else:
                    st.button(f"📷 วันที่ {chosen_day} ไม่มีรูป", disabled=True, key=f"zip_day_dis_{m_id}")
                    
            with zip_month_col:
                zip_month_data = zip_single_machine_photos(m_id, target_date_obj=photo_date_input, target_day=None)
                if zip_month_data:
                    st.download_button(label="📦 โหลดรูปทั้งเดือน", data=zip_month_data, file_name=f"Photos_{m_id}_Full_{photo_date_input.strftime('%Y_%B')}.zip", mime="application/zip", key=f"zip_month_btn_{m_id}")
                else:
                    st.button("📷 เดือนนี้ยังไม่มีรูป", disabled=True, key=f"zip_month_dis_{m_id}")
            st.divider()

        # ---- 1. แผนก CNC ----
        st.write("#### 🔹 เครื่อง CNC (9 เครื่อง)")
        cnc_col1, cnc_col2, cnc_col3 = st.columns(3)
        cnc_idx = 0
        for m_id, m_name in MACHINES.items():
            if "CNC" in m_id and "CRANE" not in m_id.upper() and "QC-" not in m_id.upper():
                with (cnc_col1 if cnc_idx % 3 == 0 else (cnc_col2 if cnc_idx % 3 == 1 else cnc_col3)):
                    render_machine_card(m_id, m_name, "CNC")
                cnc_idx += 1

        # ---- 2. แผนก GRINDING ----
        st.write("#### 🔹 เครื่องเจียรผิว GRINDING (2 เครื่อง)")
        grind_col1, grind_col2 = st.columns(2)
        grind_idx = 0
        for m_id, m_name in MACHINES.items():
            if "GRINDING" in m_id and "CUTTER" not in m_id:
                with (grind_col1 if grind_idx % 2 == 0 else grind_col2):
                    render_machine_card(m_id, m_name, "GRINDING")
                grind_idx += 1

        # ---- 3. แผนก CUTTER GRINDING ----
        st.write("#### 🔹 เครื่องลับคม CUTTER GRINDING (1 เครื่อง)")
        cutter_grind_col1, = st.columns(1)
        with cutter_grind_col1: render_machine_card("CUTTER GRINDING-01", MACHINES["CUTTER GRINDING-01"], "CUTTER GRINDING-01")

        # ---- 4. แผนกปั๊มลม COMPRESSOR ----
        st.write("#### 🔹 เครื่องปั๊มลม AIR COMPRESSOR (2 เครื่อง)")
        comp_col1, comp_col2, comp_col3 = st.columns(3)
        comp_idx = 0
        for m_id, m_name in MACHINES.items():
            if "COMP-" in m_id.upper():
                with (comp_col1 if comp_idx % 3 == 0 else (comp_col2 if comp_idx % 3 == 1 else comp_col3)):
                    render_machine_card(m_id, m_name, m_id) 
                comp_idx += 1

        # ---- 5. แผนก CRANE ----
        st.write("#### 🔹 เครน CRANE (2 แผนก)")
        crane_col1, crane_col2 = st.columns(2)
        crane_idx = 0
        for m_id, m_name in MACHINES.items():
            if "CRANE" in m_id.upper() or "Crane" in m_id:
                with (crane_col1 if crane_idx % 2 == 0 else crane_col2):
                    render_machine_card(m_id, m_name, "Crane no.1" if "no.1" in m_id else "Crane no.2") 
                crane_idx += 1

        # ---- 6. แผนก QC ----
        st.write("#### 🔹 เครื่องมือวัดคุณภาพ QC (19 เครื่องมือวัด, 2 เครื่องจักรทำงาน)")
        qc_col1, qc_col2, qc_col3 = st.columns(3)
        qc_idx = 0
        for m_id, m_name in MACHINES.items():
            if "QC-" in m_id.upper():
                with (qc_col1 if qc_idx % 3 == 0 else (qc_col2 if qc_idx % 3 == 1 else qc_col3)):
                    render_machine_card(m_id, m_name, m_id) 
                qc_idx += 1

        # ---- 7. แผนก MILLING ----
        st.write("#### 🔹 เครื่องมิลลิ่ง MILLING (4 เครื่อง)")
        mill_col1, mill_col2, mill_col3 = st.columns(3)
        mill_idx = 0
        for m_id, m_name in MACHINES.items():
            if "MILLING" in m_id:
                with (mill_col1 if mill_idx % 3 == 0 else (mill_col2 if mill_idx % 3 == 1 else mill_col3)):
                    render_machine_card(m_id, m_name, "MILLING")
                mill_idx += 1

        # ---- 8. แผนก LATHE ----
        st.write("#### 🔹 เครื่องกลึง LATHE (1 เครื่อง)")
        lathe_col1, = st.columns(1)
        with lathe_col1: render_machine_card("LATHE-01", MACHINES["LATHE-01"], "LATHE")

        # ---- 9. แผนก CUTTING ----
        st.write("#### 🔹 เครื่องตัด CUTTING (1 เครื่อง)")
        cut_col1, = st.columns(1)
        with cut_col1: render_machine_card("CUTTING-01", MACHINES["CUTTING-01"], "CUTTING")

        # ---- 10. แผนก BENDING ----
        st.write("#### 🔹 เครื่องพับ BENDING (1 เครื่อง)")
        bend_col1, = st.columns(1)
        with bend_col1: render_machine_card("BENDING-01", MACHINES["BENDING-01"], "BENDING")

        # ---- 11. แผนก MIG CO2 ----
        st.write("#### 🔹 เครื่องเชื่อม MIG CO2 (3 เครื่อง)")
        mig_col1, mig_col2, mig_col3 = st.columns(3)
        mig_idx = 0
        for m_id, m_name in MACHINES.items():
            if "MIG" in m_id:
                with (mig_col1 if mig_idx % 3 == 0 else (mig_col2 if mig_idx % 3 == 1 else mig_col3)):
                    render_machine_card(m_id, m_name, "MIG CO2")
                mig_idx += 1

        # ---- 12. แผนก ARGON ----
        st.write("#### 🔹 เครื่องเชื่อม ARGON (2 เครื่อง)")
        argon_col1, argon_col2 = st.columns(2)
        argon_idx = 0
        for m_id, m_name in MACHINES.items():
            if "ARGON" in m_id:
                with (argon_col1 if argon_idx % 2 == 0 else argon_col2):
                    render_machine_card(m_id, m_name, "ARGON")
                argon_idx += 1

        # ---- 13. แผนก WELDING ALUMINUM ----
        st.write("#### 🔹 เครื่องเชื่อมอลูมิเนียม WELDING ALUMINUM (1 เครื่อง)")
        wel_al_col1, = st.columns(1)
        with wel_al_col1: render_machine_card("WELDING_ALUMINUM-01", MACHINES["WELDING_ALUMINUM-01"], "WELDING_ALUMINUM")

        # ---- 14. แผนก BAND SAW ----
        st.write("#### 🔹 เครื่องเลื่อยสายพาน BAND SAW (3 เครื่อง)")
        saw_col1, saw_col2, saw_col3 = st.columns(3)
        saw_idx = 0
        for m_id, m_name in MACHINES.items():
            if "BAND" in m_id.upper():
                with (saw_col1 if saw_idx % 3 == 0 else (saw_col2 if saw_idx % 3 == 1 else saw_col3)):
                    render_machine_card(m_id, m_name, "BAND SAW")
                saw_idx += 1

        # ---- 15. แผนก FORKLIFT ----
        st.write("#### 🔹 รถโฟคลิฟ FORKLIFT (1 เครื่อง)")
        fork_col1, = st.columns(1)
        with fork_col1: render_machine_card("FORKLIFT-01", MACHINES["FORKLIFT-01"], "FORKLIFT")

        st.markdown("---")
        st.write("### 👑 พื้นที่ควบคุมระดับความปลอดภัยสูงสุด (สำหรับผู้บริหาร)")
        bigboss_code_input = st.text_input("🔐 กรุณากรอกรหัสผ่านผู้บริหาร เพื่อเปิดตู้นิรภัย พิมพ์คิวอาร์โค้ด และระบบล้างประวัติหลังบ้าน:", type="password", key="bigboss_secret_key_field")
        
        if bigboss_code_input == BIGBOSS_PASSWORD:
            st.success("🎯 ยืนยันสิทธิ์ สำเร็จ ปลดล็อกเรียบร้อยแล้วครับ!")
            
            with st.expander("📦 [เฉพาะผู้บริหารสูงสุด] ดาวน์โหลดไฟล์ดิบฐานข้อมูลหลัก (DATABASE BACKUP DISK)"):
                st.info("📂 ปุ่มนี้ทำหน้าที่ดึงไฟล์ประวัติการติ๊กและหมายเหตุสะสมทั้งหมดของเครื่องจักรทุกแผนกบนระบบคลาวด์กระจกเงาออกเป็นไฟล์ .csv เพื่อเก็บเป็นประวัติถาวร")
                local_cloud_backup = os.path.join(BASE_FOLDER, "gsheet_cloud_mirror.csv")
                if os.path.exists(local_cloud_backup):
                    with open(local_cloud_backup, "rb") as f_data:
                        st.download_button(
                            label="📥 ดาวน์โหลดไฟล์ฐานข้อมูลดิบกลาง (gsheet_cloud_mirror.csv)",
                            data=f_data,
                            file_name=f"Backup_Master_Database_{datetime.datetime.now().strftime('%Y_%m_%d')}.csv",
                            mime="text/csv",
                            type="primary"
                        )
                else:
                    st.caption("ℹ️ ระบบพร้อมจัดเก็บ (จะปรากฏปุ่มให้ดาวน์โหลดเมื่อช่างมีการคีย์งานส่งเข้ามาคนแรกครับ)")

            with st.expander("📸 [เฉพาะผู้บริหารสูงสุด] ดาวน์โหลดรูปภาพ PM รวมหมดทั้งโรงงาน (.zip)"):
                st.info("📦 ปุ่มนี้จะทำหน้าที่เดินสแกนกวาดรูปถ่าย PM ของทุกแผนก ทุกเครื่องจักร มารวมเป็นไฟล์ .zip ก้อนเดียวเพื่อใช้ส่งผลตรวจมาตรฐานโรงงาน")
                
                dept_target = st.selectbox("เลือกแผนกที่บอสต้องการดาวน์โหลดรูปภาพ:", [
                    "ทั้งโรงงาน", "CNC", "GRINDING", "CRANE", "COMPRESSOR", "QC", "MILLING", "MIG CO2", "ARGON", "เครื่องจักรอื่น ๆ (พับ/ตัด/กลึง/โฟคลิฟ)"
                ])
                current_boss_month = selected_date.strftime("%Y_%B")
                
                filtered_zip_data = zip_all_factory_photos_by_filter(filter_type=dept_target, target_date_obj=selected_date)
                if filtered_zip_data:
                    st.download_button(
                        label=f"📥 ดาวน์โหลดรูปภาพแผนก [{dept_target}]", 
                        data=filtered_zip_data, 
                        file_name=f"Photos_Filter_{dept_target}_{current_boss_month}.zip", 
                        mime="application/zip", 
                        type="primary"
                    )
                else:
                    st.warning(f"⚠️ ในระบบคลาวด์ตามช่วงปฏิทินที่เลือก ยังไม่มีรูปภาพบันทึกอยู่ในกลุ่มแผนก [{dept_target}] เลยครับบอส")

            with st.expander("🖨️ [เฉพาะผู้บริหารสูงสุด] เครื่องมือพิมพ์ QR Code สำหรับไปแปะหน้าเครื่องจักร"):
                sel_m = st.selectbox("เลือกเครื่องที่ต้องการพิมพ์ QR:", list(MACHINES.keys()), key="bigboss_qr_select_box")
                qr_url = f"https://pes-maintenance.streamlit.app/?id={sel_m}" 
                qr = qrcode.make(qr_url)
                buf = BytesIO()
                qr.save(buf)
                st.image(buf, caption=f"QR สำหรับแปะหน้าเครื่อง {MACHINES[sel_m]}")

            with st.expander("📦 [เฉพาะผู้บริหารสูงสุด] ตู้เซฟเก็บประวัติเอกสารย้อนหลังอัตโนมัติ (BACKUP HISTORY ARCHIVES)"):
                st.info("📂 ส่วนนี้เป็นที่รวบรวมไฟล์ Excel ประจำเดือนเก่าที่ระบบทำการคัดลอกสำรอง (Auto-Backup) เก็บไว้ให้โดยอัตโนมัติทุก ๆ สิ้นเดือน")
                backup_folder_path = os.path.join(BASE_FOLDER, "maintenance_backups")
                if os.path.exists(backup_folder_path):
                    all_backups = [f for f in os.listdir(backup_folder_path) if f.lower().endswith('.xlsx')]
                    if all_backups:
                        for b_file in sorted(all_backups):
                            b_file_path = os.path.join(backup_folder_path, b_file)
                            with open(b_file_path, "rb") as f_data:
                                st.download_button(
                                    label=f"📥 ดาวน์โหลดไฟล์สำรอง: {b_file}",
                                    data=f_data,
                                    file_name=b_file,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key=f"dl_backup_{b_file}"
                                )
                    else:
                        st.caption("ℹ️ ยังไม่มีไฟล์สำรองประวัติเดือนเก่าจัดเก็บในตู้นี้")
                else:
                    st.caption("ℹ️ ระบบกำลังเตรียมตู้เซฟ")

            with st.expander("🧹 [เฉพาะผู้บริหารสูงสุด] กล่องเครื่องมือล้างระบบภาพถ่ายและฐานข้อมูลกระจกเงา (RESET SYSTEM)"):
                st.warning("⚠️ คำเตือน: ปุ่มนี้จะทำการลบโฟลเดอร์รูปภาพหลักฐานที่ส่งทดสอบและประวัติประจักษ์กระจกเงาทั้งหมดออกไปอย่างถาวร เพื่อให้ระบบสะอาดพร้อมเปิดใช้งานจริง")
                if st.button("🚨 สั่งลบรูปภาพและฐานข้อมูลกระจกเงาทั้งหมดกริบ 100%", type="primary", key="reset_all_photos_primary_btn"):
                    target_photo_folder = os.path.join(BASE_FOLDER, "maintenance_photos")
                    local_cloud_backup = os.path.join(BASE_FOLDER, "gsheet_cloud_mirror.csv")
                    if os.path.exists(target_photo_folder): shutil.rmtree(target_photo_folder) 
                    if os.path.exists(local_cloud_backup): os.remove(local_cloud_backup)
                    st.success("🧹 ลบโฟลเดอร์รูปภาพและกระจกเงาออกไปจากระบบคลาวด์สะอาดบริสุทธิ์เรียบร้อยแล้วครับ!")
                    st.balloons()
        elif bigboss_code_input != "":
            st.error("❌ รหัสผ่าน Big Boss ไม่ถูกต้อง! ปฏิเสธสิทธิ์การพิมพ์คิวอาร์ เข้าถึงไฟล์ประวัติย้อนหลัง และปุ่มล้างระบบ")

    elif password_input != "": 
        st.error("❌ รหัสผ่านไม่ถูกต้อง ไม่พบสิทธิ์เข้าใช้งานระบบตามรหัสนี้ครับ")
